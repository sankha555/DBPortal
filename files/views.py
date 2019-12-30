from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import requests
from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from .forms import FileUploadForm, SearchUIDForm, MailContentForm, AddDberForm, SearchCityStaffForm
import xlrd, xlsxwriter
import openpyxl
from django.conf import settings
from .models import File
from users.models import Profile


def home(request):
    return render(request, 'files/home.html')

@login_required
def upload_file(request):

    if File.objects.filter(staff = request.user.profile).exists():
        return HttpResponse("You have already uploaded your file, Update the previous file.")
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.save(commit=False)
            file.save()

            #file = request.FILES['filename'].temporary_file_path
            document = file.doc
            wb = xlrd.open_workbook(document.path)
            sheet = wb.sheet_by_index(0)
            sheet.cell_value(0, 0)

            rows = sheet.nrows
            cols = sheet.ncols

            for i in range(2,rows):
                dber_uid = sheet.cell_value(i, 0)
                dber_name = sheet.cell_value(i, 1)
                dber_dob = sheet.cell_value(i, 2)
                dber_gender = sheet.cell_value(i, 3)
                dber_city = sheet.cell_value(i, 4)
                dber_state = sheet.cell_value(i, 5)

                new_dber = Profile.objects.get_or_create(uid = dber_uid)
                new_dber.name = dber_name
                new_dber.dob = dber_dob
                new_dber.gender = dber_gender
                new_dber.city = dber_city
                new_dber.state = dber_state
                new_dber.city_circle = file.staff.city_circle

                new_dber.save()

            return redirect('home')
    else:
        form = FileUploadForm()

    return render(request, 'files/upload_file.html', {
        'form': form
    })

def search_dber(request):

    if request.method == 'POST':
        form = SearchUIDForm(request.POST)
        if form.is_valid():
            search_uid = form.cleaned_data['uid']

            file = get_object_or_404(File, staff = request.user.profile)
            document = file.doc
            wb = xlrd.open_workbook(document.path)
            sheet = wb.sheet_by_index(0)

            rows = sheet.nrows
            cols = sheet.ncols

            flag = 0

            for i in range(2,rows):

                if search_uid == sheet.cell_value(i, 0):
                    flag = 1
                    dber = get_object_or_404(Profile, uid = search_uid)
                    return render(request, "files/search_results.html", {'dber':dber})

            if flag == 0:
                return render(request, "files/failed_search.html")
    else:
        form = SearchUIDForm()

    return render(request, 'files/search_dber.html', {
        'form': form
    })

def search_uid_normal(request):

    if request.method == 'POST':
        form = SearchUIDForm(request.POST)
        if form.is_valid():
            search_uid = form.cleaned_data['uid']

            dber = get_object_or_404(Profile, uid = search_uid)
            return render(request, "files/search_results.html", {'dber':dber})

    else:
        form = SearchUIDForm()

    return render(request, 'files/search_dber.html', {
        'form': form
    })

@login_required
def update_dber(request, pk):

    dber = get_object_or_404(Profile, pk = pk)
    if request.method == 'POST':
        p_form = DberUpdateForm(request.POST,
                                    instance=dber)

        if p_form.is_valid():
            dber.save()
            p_form.save()

            # File updation begins

            file = get_object_or_404(File, staff = request.user.profile)
            document = file.doc
            wb = openpyxl.load_workbook(document.path)
            sheet = wb.get_sheet(0)

            rows = sheet.max_row

            flag = 0
            for i in range(3,rows+1):

                cell_name = f'A{i}'
                if dber.uid == sheet[cell_name].value:
                    flag = 1

                    list = [dber.uid, dber.name, dber.dob, dber.gender, dber.city, dber.state]
                    for j in range(2, 7):
                        cell = sheet.cell(row = i, column = j)
                        cell.value = list[j-1]

                    break

            if flag == 0:
                return render(request, "files/failed_search.html")

            wb.save(document.path)
            file.save()
            #File Updation Ends

            return redirect('home')

    else:
        p_form = DberUpdateForm(instance=dber)


    context = {
        'p_form' : p_form
    }
    return render(request, 'files/update_dber.html', context)

@login_required
def delete_dber(request, pk):

    dber = get_object_or_404(Profile, pk = pk)

    # File updation begins
    file = get_object_or_404(File, staff = request.user.profile)
    document = file.doc
    wb = openpyxl.load_workbook(document.path)
    sheet = wb.get_sheet(0)

    rows = sheet.max_row

    flag = 0
    for i in range(3,rows+1):
        cell_name = f'A{i}'
        if dber.uid == sheet[cell_name].value:
            flag = 1
            sheet.delete_rows(i, 1)
            break

    Profile.objects.remove(dber)

    wb.save(document.path)
    file.save()
    #File Updation Ends

    return redirect('home')

@login_required
def add_dber(request):

    if request.method == 'POST':
        form = AddDberForm(request.POST)

        if form.is_valid():
            form.save()
            dber_uid = form.cleaned_data['uid']

            #new dber created in database
            new_dber = Profile.objects.get_or_create(uid = dber_uid)
            new_dber.name = form.cleaned_data['name']
            new_dber.dob = form.cleaned_data['dob']
            new_dber.gender = form.cleaned_data['gender']
            new_dber.city = form.cleaned_data['city']
            new_dber.state = form.cleaned_data['state']
            new_dber.city_circle = request.user.profile.city_circle

            new_dber.save()

            # File updation begins
            file = get_object_or_404(File, staff = request.user.profile)
            document = file.doc
            wb = openpyxl.load_workbook(document.path)
            sheet = wb.get_sheet(0)

            rows = sheet.max_row
            cols = sheet.max_column

            sheet.insert_rows(3) #new dber insterted at top of sheet

            list = [new_dber.uid, new_dber.name, new_dber.dob, new_dber.gender, new_dber.city, new_dber.state]
            for j in range(1, 7):
                cell = sheet.cell(row = 3, column = j)
                cell.value = list[j-1]


            wb.save(document.path)
            file.save()
            #File Updation Ends

            return redirect('home')

    else:
        form = AddDberForm()

    return render(request, 'files/add_dber.html', {'form':form})

@login_required
def send_mail(request, pk):

    dber = get_object_or_404(Profile, pk = pk)
    if request.method == 'POST':
        form = MailContentForm(request.POST)

        if form.is_valid():

            subject = form.cleaned_data['subject']
            content = form.cleaned_data['content']

            mail = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mail.ehlo()
            mail.starttls()
            mail.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            message = f'**** SUBJECT : {subject} **** \n{content}\n\nRegards,\n{request.user.profile.name}'

            email = dber.user.email
            try:
                mail.sendmail(settings.EMAIL_HOST_USER, email, message)
            except:
                pass

            mail.close()

    else:
        form = MailContentForm()

    return render(request, 'files/mail_content.html', {'form':form})

@login_required
def send_mass_mails(request):

    if request.method == 'POST':
        form = MailContentForm(request.POST)

        if form.is_valid():

            subject = form.cleaned_data['subject']
            content = form.cleaned_data['content']

            mail = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mail.ehlo()
            mail.starttls()
            mail.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            message = f'**** SUBJECT : {subject} **** \n{content}\n\nRegards,\n{request.user.profile.name}'

            users = User.objects.all(profile)
            for user in users:
                if (user.profile.city_circle == request.user.profile.city_circle) and (user.profile.qualifier == 0):

                    email = user.email
                    try:
                        mail.sendmail(settings.EMAIL_HOST_USER, email, message)
                    except:
                        pass

            mail.close()

    else:
        form = MailContentForm()

    return render(request, 'files/mail_content.html', {'form':form})

@login_required
def send_city_staff_mail(request):

    if request.method == 'POST':
        form = SearchCityStaffForm(request.POST)

        if form.is_valid():

            city = form.cleaned_data['city']

            flag = 0
            staff_pool = Profile.objects.exclude(qualifier=0)

            for staff_user in staff_pool:
                if staff_user.city == city:
                    flag = 1
                    return redirect('send_mail', staff_user.id)
                    break

            if flag == 0:
                return render(request, 'templates/failed_search.html')

    else:


        form = SearchCityStaffForm()

    return render(request, 'files/search_results.html', {'form':form})

@login_required
def link_account(request, pk):
    profile = get_object_or_404(Profile, pk = pk)
    user = request.user

    profile.user = user
    profile.save()
    user.save()

    return redirect('home')

@login_required
def search_city_staff(request):

    if request.method == "POST":
        form = SearchCityStaffForm(request.POST)

        if form.is_valid():
            search_city = form.cleaned_data['city']

            if Profile.objects.filter(qualifier = 1, city = search_city).exists():
                staff = Profile.objects.get(qualifier = 1, city = search_city)
                return redirect('send_mail', staff.id)
            else:
                return redirect('search_city_staff')

    else:
        form = SearchCityStaffForm()

    return render(request, 'files/search_city_staff.html', {'form':form})
